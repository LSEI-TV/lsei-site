# Convertit les fichiers de classement par saison (YYYY-YYYY.xlsx) + participantsView.xlsx
# en results-YYYY-YYYY.json (format du site), avec CALENDRIER COMPLET depuis les légendes.
import sys, json, re, unicodedata, hashlib
from collections import defaultdict
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
OUT = '../src/data/billard'
CUES = ['results-2025-2026', 'results-2024-2025', 'results-2023-2024']
SEASONS = ['2015-2016', '2016-2017', '2017-2018', '2018-2019', '2019-2020', '2021-2022', '2022-2023']
ANN2SEASON = {'3': '2016-2017', '4': '2017-2018', '5': '2018-2019', '6': '2019-2020', '8': '2021-2022', '9': '2022-2023'}
ROUND = {'seizieme de finale': 'last 32', 'huitieme de finale': 'last 16', 'quart de finale': 'quarter final',
         'demi-finale': 'semi final', 'finale': 'final'}

def norm(s):
    s = unicodedata.normalize('NFD', str(s or '')).encode('ascii', 'ignore').decode()
    return ''.join(c for c in s.lower() if c.isalnum() or c == ' ').strip()
def namekey(name): return ' '.join(sorted(norm(name).split()))
def slugify(s):
    s = unicodedata.normalize('NFD', str(s or '')).encode('ascii', 'ignore').decode().lower()
    return re.sub(r'^-|-$', '', re.sub(r'[^a-z0-9]+', '-', s))
def reorder(name):
    toks = str(name or '').split()
    if not toks: return ''
    up = [t for t in toks if t.isupper() and len(t) > 1]
    if up and len(up) < len(toks):
        return ' '.join(t.capitalize() for t in [t for t in toks if t not in up] + [t for t in toks if t in up])
    return ' '.join(toks[-1:] + toks[:-1])

def level_kind(name):
    n = str(name or '').strip()
    if re.search(r'championnat[s]? de france', n, re.I) or re.fullmatch(r'cdf', n, re.I):
        return ('Championnat de France', 'france')
    m = (re.search(r'n[°o]\s*0*(\d+)', n, re.I) or re.search(r'\btn\s*0*(\d+)', n, re.I)
         or re.search(r'(\d+)\s*(?:er|eme|ème|nd|e)?\s+tournoi', n, re.I))
    return (('TN' + m.group(1)) if m else n[:22], 'tn')
def city_of(name):
    n = name.split(':', 1)[-1]
    for pat in [r'saison\s*20\d\d\s*-\s*20\d\d', r'20\d\d\s*-\s*20\d\d', r'20\d\d', r'n[°o]\s*\d+',
                r'\d+\s*(?:er|eme|ème|nd|e)?\s+tournoi national', r'tournoi national',
                r'championnat[s]? de france', r'(?:et\s+)?coupe de france.*', r'-?\s*principal',
                r'\bb{1,2}m\b', r'\bffb\b', r'\bet\b', r'indiv\w*', r'hors\s+r[ée]gionaux', r'r[ée]gionaux',
                r'individuelle', r'doublette']:
        n = re.sub(pat, ' ', n, flags=re.I)
    n = re.sub(r'\s+', ' ', n).strip(' -,')
    n = re.sub(r'^(a|à|aux|au|de|le|la)\s+', '', n, flags=re.I).strip(' -')
    return n.title()

# --- réconciliation nom -> id Cuescore ---
cues_id, cues_name = {}, {}
for f in CUES:
    for p in json.load(open(f'{OUT}/{f}.json', encoding='utf-8'))['players']:
        k = namekey(p['name']); cues_id.setdefault(k, p['id']); cues_name.setdefault(k, p['name'])
# Réconciliation par NOM (clé order-independent) valable sur toutes les saisons :
# id Cuescore si connu ; sinon chiffres de la licence ; sinon (pas de licence,
# ex. fichier 2015-2016) id synthétique stable dérivé du nom. Un même joueur
# garde le même id d'une saison à l'autre, licence ou pas.
NAME2ID, NAME2NAME = {}, {}
def resolve(lic, raw):
    disp = reorder(raw); k = namekey(disp)
    if not k: return {'id': 0, 'name': disp}
    if k in cues_id: return {'id': cues_id[k], 'name': cues_name[k]}
    if k in NAME2ID: return {'id': NAME2ID[k], 'name': NAME2NAME[k]}
    d = re.sub(r'\D', '', str(lic or ''))
    NAME2ID[k] = int(d) if d else 800000000 + int(hashlib.md5(k.encode()).hexdigest()[:7], 16)
    NAME2NAME[k] = disp
    return {'id': NAME2ID[k], 'name': disp}

# --- lecture d'un fichier de classement : légende (villes) + roster + points/étape ---
LEG_RE = re.compile(r'^(?:ev|tn)\s*\d+\s*:\s*(.+)$', re.I)
def _num(v):
    try: return int(str(v).strip())
    except: return None
def read_season_file(season):
    ws = openpyxl.load_workbook(f'{season}.xlsx', read_only=True, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == 'Rang')
    hdr = [str(c).strip() if c is not None else '' for c in rows[hi]]
    # Légende (dans l'ordre) — préfixe « Ev N : » ou « TN N : ». Niveau/ville lus dans le libellé.
    legend_all = []
    for i in range(hi):
        m = LEG_RE.match(str(rows[i][0] or '').strip())
        if not m: continue
        desc = m.group(1); carried = bool(re.search(r'classement\s+tn', desc, re.I))  # points reportés COVID
        lvl, kind = level_kind(desc)
        legend_all.append({'level': lvl, 'kind': kind, 'city': city_of(desc), 'carried': carried})
    legend = [e for e in legend_all if not e['carried']]
    # Colonnes : par nom d'en-tête (la colonne Licence n'existe pas dans tous les fichiers).
    idx = lambda n: hdr.index(n) if n in hdr else None
    ci_rang, ci_lic, ci_name, ci_pts, ci_ga = idx('Rang'), idx('Licence'), idx('Nom Prénom'), idx('Total points'), idx('GA cumulé')
    ev_cols = [j for j in range((ci_ga if ci_ga is not None else ci_pts) + 1, len(hdr)) if hdr[j]]
    # Chaque colonne d'événement ↔ event de la légende, dans l'ordre.
    col_lvl = []
    for k, j in enumerate(ev_cols):
        if k < len(legend_all): col_lvl.append((legend_all[k]['level'], legend_all[k]['carried']))
        else: col_lvl.append((level_kind(hdr[j])[0], False))
    players = []
    for r in rows[hi + 1:]:
        if not r or ci_rang >= len(r) or not str(r[ci_rang]).strip().isdigit(): continue
        name = str(r[ci_name] or '').strip()
        if not name: continue
        ev = {}
        for k, j in enumerate(ev_cols):
            lvl, carried = col_lvl[k]
            if carried or j >= len(r): continue
            v = _num(r[j])
            if v is not None: ev[lvl] = v
        players.append({'lic': str(r[ci_lic]).strip() if ci_lic is not None and ci_lic < len(r) else '',
                        'raw': name, 'rank': int(str(r[ci_rang]).strip()),
                        'points': _num(r[ci_pts]) or 0, 'ev': ev})
    return legend, players

# --- matchs participantsView par saison ---
wbp = openpyxl.load_workbook('participantsView.xlsx', read_only=True, data_only=True)
prows = list(wbp.active.iter_rows(values_only=True)); ph = prows[0]; pc = lambda n: ph.index(n)
matches_by_season = defaultdict(list)
for r in prows[1:]:
    season = ANN2SEASON.get(str(r[pc('annee_sportive_id')]))
    if not season: continue
    try: sA = int(r[pc('manches_gagnees_joueur1')]); sB = int(r[pc('manches_gagnees_joueur2')])
    except: continue
    a = resolve(r[pc('licence_joueur1')], r[pc('nom_joueur1')]); b = resolve(r[pc('licence_joueur2')], r[pc('nom_joueur2')])
    if a['id'] == b['id']: continue
    lvl, kind = level_kind(str(r[pc('titre_court_evenement')] or ''))
    dt = str(r[pc('horaire_debut')] or '')[:10]
    matches_by_season[season].append({'lvl': lvl, 'kind': kind, 'date': dt if re.match(r'\d{4}-\d\d-\d\d', dt) else '',
        'aId': a['id'], 'bId': b['id'], 'aName': a['name'], 'bName': b['name'], 'sA': sA, 'sB': sB,
        'round': ROUND.get(str(r[pc('libelle')] or '').strip().lower(), str(r[pc('libelle')] or ''))})

def lvl_num(l):
    m = re.search(r'(\d+)', l); return (1, 0) if l.lower().startswith('champ') else (0, int(m.group(1)) if m else 50)

def build_season(season, legend, roster):
    ms = matches_by_season.get(season, [])
    if season == '2021-2022':  # le « TN7 » d'Agen est en réalité le BM N°7 / Championnat de France
        for m in ms:
            if m['lvl'] == 'TN7': m['lvl'], m['kind'] = 'Championnat de France', 'france'
    # tournois vus dans les matchs
    tinfo = {}
    for m in ms: tinfo.setdefault(m['lvl'], {'kind': m['kind'], 'dates': []})['dates'].append(m['date'])
    # fusion avec la légende (calendrier complet + villes)
    city_by_level = {e['level']: e['city'] for e in legend}
    all_levels = list(dict.fromkeys([e['level'] for e in legend] + list(tinfo.keys())))
    tours, tid_of = [], {}
    for i, lvl in enumerate(sorted(all_levels, key=lvl_num)):
        tid = f'{season}-{i+1}'; tid_of[lvl] = tid
        info = tinfo.get(lvl); ds = [d for d in (info['dates'] if info else []) if d]
        kind = info['kind'] if info else next((e['kind'] for e in legend if e['level'] == lvl), 'tn')
        tours.append({'id': tid, 'name': lvl, 'level': lvl, 'city': city_by_level.get(lvl, ''),
                      'date': min(ds) if ds else '', 'kind': kind})
    # joueurs (classement) + points/étape (parcours) ; meilleur score/étape = vainqueur potentiel
    players = {}; best_pts = {}
    for info in roster:
        r = resolve(info['lic'], info['raw'])
        evs = [{'tid': tid_of[l], 'level': l, 'points': pts} for l, pts in info['ev'].items() if l in tid_of]
        players[r['id']] = {'id': r['id'], 'name': r['name'], 'slug': slugify(r['name']), 'country': 'FRA',
                            'rank': info['rank'], 'points': info['points'], 'played': 0, 'wins': 0, 'losses': 0,
                            'winPct': 0, 'pf': 0, 'pa': 0, 'diff': 0, 'tourns': 0, 'events': evs}
        for l, pts in info['ev'].items():
            if pts is not None and (l not in best_pts or pts > best_pts[l][0]):
                best_pts[l] = (pts, r['id'], r['name'])
    outms = []; tset = defaultdict(set)
    for m in ms:
        tid = tid_of.get(m['lvl'])
        outms.append({'tid': tid, 'date': m['date'], 'aId': m['aId'], 'bId': m['bId'], 'aName': m['aName'],
                      'bName': m['bName'], 'sA': m['sA'], 'sB': m['sB'], 'round': m['round']})
        for pid, my, opp, nm in [(m['aId'], m['sA'], m['sB'], m['aName']), (m['bId'], m['sB'], m['sA'], m['bName'])]:
            if pid not in players:
                players[pid] = {'id': pid, 'name': nm, 'slug': slugify(nm), 'country': 'FRA', 'rank': 0, 'points': 0,
                                'played': 0, 'wins': 0, 'losses': 0, 'winPct': 0, 'pf': 0, 'pa': 0, 'diff': 0, 'tourns': 0, 'events': []}
            p = players[pid]; p['played'] += 1; p['pf'] += my; p['pa'] += opp
            p['wins' if my > opp else 'losses'] += 1; tset[pid].add(tid)
    for pid, p in players.items():
        p['diff'] = p['pf'] - p['pa']; p['tourns'] = len(tset[pid] | {e['tid'] for e in p['events']})
        p['winPct'] = round(p['wins'] / p['played'] * 100) if p['played'] else 0
    # Vainqueur par tournoi : finale jouée (matchs) sinon meilleur score/étape (points).
    for t in tours:
        fin = next((m for m in outms if m['tid'] == t['id'] and (m['round'] or '').lower() in ('final', 'grand final')), None)
        if fin:
            t['winner'] = {'id': fin['aId'], 'name': fin['aName']} if fin['sA'] > fin['sB'] else {'id': fin['bId'], 'name': fin['bName']}
        elif t['level'] in best_pts:
            _, wid, wname = best_pts[t['level']]; t['winner'] = {'id': wid, 'name': wname}
    if season == '2021-2022':   # points FFB reconstruits depuis le placement en tableau
        LOSS = {'final': 364, 'semi final': 292, 'quarter final': 224, 'last 16': 160, 'last 32': 100}
        RK = {'last 32': 1, 'last 16': 2, 'quarter final': 3, 'semi final': 4, 'final': 5}
        deep = {}   # (pid, tid) -> (profondeur, round, gagné)
        for m in outms:
            for pid, my, opp in [(m['aId'], m['sA'], m['sB']), (m['bId'], m['sB'], m['sA'])]:
                rd = (m['round'] or '').lower(); key = (pid, m['tid'])
                if key not in deep or RK.get(rd, 0) > deep[key][0]: deep[key] = (RK.get(rd, 0), rd, my > opp)
        lvl_of = {t['id']: t['level'] for t in tours}; order = [t['id'] for t in tours]
        pev = defaultdict(list)
        for pid in players: players[pid]['points'] = 0
        for (pid, tid), (_, rd, won) in deep.items():
            pts = 440 if (rd == 'final' and won) else LOSS.get(rd, 100)
            players[pid]['points'] += pts
            pev[pid].append({'tid': tid, 'level': lvl_of.get(tid, tid), 'points': pts})
        for pid, p in players.items():
            p['events'] = sorted(pev.get(pid, []), key=lambda e: order.index(e['tid']))
            p['tourns'] = len(pev.get(pid, []))
        for i, p in enumerate(sorted(players.values(), key=lambda p: (-p['points'], -p['wins']))): p['rank'] = i + 1
    plist = sorted(players.values(), key=lambda p: (p['rank'] if p['rank'] else 9999, -p['points']))
    return {'competition': 'Blackball Master', 'season': season.replace('-', '/'),
            'tournaments': tours, 'players': plist, 'matches': outms}

results = {}
for season in SEASONS:
    legend, roster = read_season_file(season)
    if season == '2021-2022':   # le fichier est le classement 2020/2021 reporté (COVID) :
        roster = []             # inutilisable pour les points -> saison reconstruite depuis les matchs
    results[season] = build_season(season, legend, roster)

for season, R in results.items():
    open(f'{OUT}/results-{season}.json', 'w', encoding='utf-8').write(json.dumps(R, ensure_ascii=False))
    champ = min(R['players'], key=lambda p: p['rank'] if p['rank'] else 9999)
    cities = sum(1 for t in R['tournaments'] if t['city'])
    print(f"{season}: {len(R['players'])} j · {len(R['tournaments'])} tournois ({cities} avec ville) · {len(R['matches'])} matchs "
          f"| Champion: {champ['name']} ({champ['points']}pts)")
